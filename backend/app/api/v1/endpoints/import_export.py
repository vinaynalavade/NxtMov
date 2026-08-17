import os
import csv
import uuid
import io
import logging
import openpyxl
from io import BytesIO
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from app.core.database import get_db
from app.core.tenant import get_current_tenant, TenantContext
from app.core.permissions import require_any_permission, Permission
from app.models.company import Company, Contact, ContactStatus
from app.models.candidate import Candidate, CandidateStatus
from app.schemas.import_export import (
    ImportPreviewResponse, ImportPreviewRow, ColumnMappingInfo,
    ImportConfirmRequest, ImportResultResponse
)

logger = logging.getLogger(__name__)

router = APIRouter()

UPLOAD_TMP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), "tmp_uploads")
os.makedirs(UPLOAD_TMP_DIR, exist_ok=True)

HR_SYNONYMS = {
    "name": [
        "contact name", "hr name", "recruiter", "recruiter name", "contact person", "hr contact",
        "talent acquisition", "ta", "ta contact", "full name", "person name", "contact", "name", "hr"
    ],
    "company_name": [
        "company name", "company", "organisation", "organization", "employer", "client",
        "firm name", "firm", "corporate name", "org name", "org"
    ],
    "designation": [
        "designation", "role", "position", "job title", "title", "hr designation"
    ],
    "phone": [
        "mobile", "phone", "phone number", "mobile number", "contact number", "mobile no",
        "phone no", "contact no", "telephone", "cell", "cell number", "hr phone", "phone #"
    ],
    "email": [
        "email", "email id", "email address", "e-mail", "mail", "contact email", "hr email"
    ],
    "location": [
        "location", "city", "work location", "address", "state", "office location"
    ],
    "linkedin_url": [
        "linkedin", "linkedin url", "linkedin profile", "linkedin link", "profile link"
    ],
    "notes": [
        "notes", "remarks", "comments", "description", "details"
    ]
}

CANDIDATE_SYNONYMS = {
    "name": [
        "student name", "candidate name", "full name", "applicant name", "name", "person name", "candidate"
    ],
    "email": [
        "email", "email address", "email id", "contact email", "mail", "candidate email"
    ],
    "phone": [
        "phone", "mobile", "contact no", "mobile no", "phone number", "contact number", "cell", "telephone"
    ],
    "current_title": [
        "current title", "designation", "role", "current role", "job title", "title", "profile"
    ],
    "primary_skills": [
        "technical skills", "skills", "primary skills", "key skills", "technologies", "tech stack"
    ],
    "experience_years": [
        "experience", "total exp", "relevant exp", "exp (yrs)", "exp", "years of experience", "yoe", "experience years"
    ],
    "location": [
        "location", "city", "current location", "address", "state"
    ],
    "current_company": [
        "current company", "employer", "organization", "company", "current employer"
    ],
    "notice_period_days": [
        "notice period", "notice period (days)", "notice"
    ],
    "expected_salary": [
        "expected ctc", "expected salary", "exp ctc"
    ],
    "current_salary": [
        "current ctc", "current salary"
    ],
    "notes": [
        "notes", "remarks", "comments", "summary", "about"
    ]
}

def auto_suggest_mapping(headers: List[str], import_type: str = "HR_CONTACTS") -> Tuple[Dict[str, str], List[ColumnMappingInfo]]:
    synonyms_dict = CANDIDATE_SYNONYMS if import_type == "CANDIDATES" else HR_SYNONYMS
    mapping = {}
    infos: List[ColumnMappingInfo] = []

    for header in headers:
        header_clean = header.strip().lower()
        matched = False
        matched_field = "ignore"
        confidence = "UNMAPPED"
        reason = None

        # Check exact synonyms
        for std_field, synonyms in synonyms_dict.items():
            if header_clean in synonyms:
                matched_field = std_field
                matched = True

                # Flag generic names as AMBIGUOUS if header is ambiguous
                if header_clean in ["name", "contact", "org", "details"]:
                    confidence = "AMBIGUOUS"
                    reason = f"Header '{header}' can refer to multiple fields. Please verify mapping."
                else:
                    confidence = "HIGH"
                    reason = f"Exact synonym match for '{header_clean}'."
                break

        # Check substring match
        if not matched:
            for std_field, synonyms in synonyms_dict.items():
                if any(s in header_clean for s in synonyms if len(s) > 3):
                    matched_field = std_field
                    matched = True
                    confidence = "AMBIGUOUS" if len(header_clean) < 6 else "HIGH"
                    reason = f"Matched keyword in '{header}'."
                    break

        if not matched:
            matched_field = "ignore"
            confidence = "UNMAPPED"
            reason = "No matching field found automatically."

        mapping[header] = matched_field
        infos.append(ColumnMappingInfo(
            source_header=header,
            target_field=matched_field,
            confidence=confidence,
            reason=reason
        ))

    return mapping, infos

def normalize_phone(phone_str: Optional[Any]) -> Optional[str]:
    if not phone_str:
        return None
    cleaned = "".join(c for c in str(phone_str) if c.isdigit())
    if len(cleaned) >= 10:
        return cleaned[-10:]
    return cleaned if len(cleaned) >= 7 else None

def normalize_email(email_str: Optional[Any]) -> Optional[str]:
    if not email_str:
        return None
    val = str(email_str).strip().lower()
    return val if "@" in val and "." in val else None

def normalize_company_name(comp_name: Optional[Any]) -> Optional[str]:
    if comp_name is None:
        return None
    name = str(comp_name).strip().lower()
    if not name:
        return None
    suffixes = ["pvt. ltd.", "pvt ltd", "ltd.", "ltd", "limited", "inc.", "inc", "corp.", "corp", "llp", "co."]
    for suff in suffixes:
        if name.endswith(" " + suff):
            name = name[:-len(" " + suff)].strip()
            break
    name = "".join(c for c in name if c.isalnum() or c.isspace())
    res = " ".join(name.split())
    return res if res else None

def normalize_person_name(name_str: Optional[Any]) -> Optional[str]:
    if name_str is None:
        return None
    res = " ".join(str(name_str).strip().lower().split())
    return res if res else None

def format_clean_name(name_str: Optional[Any]) -> Optional[str]:
    if name_str is None:
        return None
    res = " ".join(str(name_str).strip().title().split())
    return res if res else None

@router.post("/preview", response_model=ImportPreviewResponse, summary="Upload & Preview Excel/CSV File")
async def preview_import(
    file: UploadFile = File(...),
    import_type: str = Query("HR_CONTACTS", alias="import_type"),
    sheet_name: Optional[str] = Query(None, alias="sheet_name"),
    ctx: TenantContext = Depends(get_current_tenant),
    db: Session = Depends(get_db)
):
    original_filename = file.filename or "upload.xlsx"
    ext = os.path.splitext(original_filename)[1].lower()
    if ext not in [".xlsx", ".csv"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Please upload a valid .xlsx or .csv file."
        )

    content = await file.read()
    file_size_bytes = len(content)
    if file_size_bytes > 15 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File size exceeds 15MB limit.")

    file_token = f"{uuid.uuid4().hex}{ext}"
    saved_path = os.path.join(UPLOAD_TMP_DIR, file_token)
    with open(saved_path, "wb") as f:
        f.write(content)

    headers: List[str] = []
    rows_data: List[Dict[str, Any]] = []
    sheets: List[str] = ["Sheet1"]
    selected_sheet = sheet_name or "Sheet1"

    try:
        try:
            if ext == ".xlsx":
                wb = openpyxl.load_workbook(saved_path, data_only=True)
                sheets = wb.sheetnames
                selected_sheet = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
                sheet = wb[selected_sheet]
                all_rows = list(sheet.iter_rows(values_only=True))
                if all_rows:
                    raw_header_row = all_rows[0]
                    headers = [
                        str(cell).strip() if (cell is not None and str(cell).strip() != "") else f"Unnamed_Col_{col_idx + 1}"
                        for col_idx, cell in enumerate(raw_header_row)
                    ]
                    for row_idx, row in enumerate(all_rows[1:], start=2):
                        if row and any(val is not None and str(val).strip() != "" for val in row):
                            row_dict = {}
                            for col_idx, val in enumerate(row):
                                if col_idx < len(headers):
                                    row_dict[headers[col_idx]] = str(val).strip() if val is not None else ""
                            rows_data.append({"row_number": row_idx, "data": row_dict})
            else:  # .csv
                text_content = content.decode("utf-8-sig", errors="ignore")
                reader = csv.reader(io.StringIO(text_content))
                all_rows = list(reader)
                if all_rows:
                    raw_header_row = all_rows[0]
                    headers = [
                        str(cell).strip() if str(cell).strip() != "" else f"Unnamed_Col_{col_idx + 1}"
                        for col_idx, cell in enumerate(raw_header_row)
                    ]
                    for row_idx, row in enumerate(all_rows[1:], start=2):
                        if row and any(str(val).strip() != "" for val in row):
                            row_dict = {}
                            for col_idx, val in enumerate(row):
                                if col_idx < len(headers):
                                    row_dict[headers[col_idx]] = str(val).strip()
                            rows_data.append({"row_number": row_idx, "data": row_dict})
        except HTTPException:
            raise
        except Exception as parse_err:
            logger.error(f"Spreadsheet parsing error for '{original_filename}': {parse_err}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unable to read or parse spreadsheet file '{original_filename}'. Detail: {str(parse_err)}"
            )

        if not headers or not rows_data:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded spreadsheet is empty or contains no readable data rows.")

        suggested_mappings, column_mapping_infos = auto_suggest_mapping(headers, import_type)

        # Pre-fetch existing workspace records for multi-signal duplicate checking
        existing_email_map: Dict[str, Any] = {}
        existing_phone_map: Dict[str, Any] = {}
        existing_name_company_map: Dict[Tuple[str, str], Any] = {}

        if import_type == "CANDIDATES":
            cand_records = db.query(Candidate).filter(Candidate.organization_id == ctx.organization.id).all()
            for c in cand_records:
                if c.email:
                    existing_email_map[c.email.strip().lower()] = c
                norm_p = normalize_phone(c.phone)
                if norm_p:
                    existing_phone_map[norm_p] = c
        else:
            contact_records = db.query(Contact).filter(Contact.organization_id == ctx.organization.id).all()
            for c in contact_records:
                if c.email:
                    existing_email_map[c.email.strip().lower()] = c
                norm_p = normalize_phone(c.phone)
                if norm_p:
                    existing_phone_map[norm_p] = c
                
                norm_name = normalize_person_name(c.name)
                norm_comp = normalize_company_name(c.company.name if c.company else None)
                if norm_name and norm_comp:
                    existing_name_company_map[(norm_name, norm_comp)] = c

        preview_rows: List[ImportPreviewRow] = []
        new_count = 0
        exact_duplicates = 0
        possible_duplicates = 0
        invalid_count = 0

        for item in rows_data:
            r_num = item["row_number"]
            r_data = item["data"]

            mapped_fields = {}
            for original_col, mapped_target in suggested_mappings.items():
                if mapped_target != "ignore" and original_col in r_data:
                    mapped_fields[mapped_target] = r_data[original_col]

            raw_person_name = mapped_fields.get("name")
            person_name = format_clean_name(raw_person_name)
            raw_company_name = mapped_fields.get("company_name")
            company_name = raw_company_name.strip() if (raw_company_name and raw_company_name.strip()) else None
            raw_email = mapped_fields.get("email")
            email = normalize_email(raw_email) or (raw_email.strip() if (raw_email and raw_email.strip()) else None)
            phone = mapped_fields.get("phone") or None
            designation = mapped_fields.get("designation") or None
            location = mapped_fields.get("location") or None
            linkedin_url = mapped_fields.get("linkedin_url") or None
            notes = mapped_fields.get("notes") or None
            skills = mapped_fields.get("primary_skills") or None

            status_flag = "NEW"
            duplicate_reason = None
            issue_details = None

            if not person_name and not email and not phone:
                status_flag = "INVALID"
                issue_details = "Row missing required contact name, email, or phone."
                invalid_count += 1
            else:
                clean_email = email.strip().lower() if email else None
                norm_phone = normalize_phone(phone)
                norm_name = normalize_person_name(person_name)
                norm_comp = normalize_company_name(company_name)

                # Signal 1: Exact Email
                if clean_email and clean_email in existing_email_map:
                    status_flag = "EXACT_DUPLICATE"
                    duplicate_reason = f"Exact Email match ({clean_email})"
                    exact_duplicates += 1
                # Signal 2: Exact Phone
                elif norm_phone and norm_phone in existing_phone_map:
                    status_flag = "EXACT_DUPLICATE"
                    duplicate_reason = f"Exact 10-digit Phone match ({norm_phone})"
                    exact_duplicates += 1
                # Signal 3: Same Name + Same Company
                elif norm_name and norm_comp and (norm_name, norm_comp) in existing_name_company_map:
                    status_flag = "POSSIBLE_DUPLICATE"
                    duplicate_reason = f"Matching HR Name & Company ({person_name} at {company_name})"
                    possible_duplicates += 1
                else:
                    status_flag = "NEW"
                    new_count += 1

            preview_rows.append(ImportPreviewRow(
                row_number=r_num,
                name=person_name,
                company_name=company_name,
                phone=phone,
                email=email,
                designation=designation,
                location=location,
                linkedin_url=linkedin_url,
                notes=notes,
                skills=skills,
                status_flag=status_flag,
                duplicate_reason=duplicate_reason,
                issue_details=issue_details
            ))

        return ImportPreviewResponse(
            file_name=file_token,
            original_filename=original_filename,
            file_size_bytes=file_size_bytes,
            import_type=import_type,
            total_rows=len(rows_data),
            selected_sheet=selected_sheet,
            sheets=sheets,
            detected_headers=headers,
            column_mappings=column_mapping_infos,
            suggested_mappings=suggested_mappings,
            preview_rows=preview_rows,
            summary_stats={
                "new_count": new_count,
                "exact_duplicates": exact_duplicates,
                "possible_duplicates": possible_duplicates,
                "invalid_count": invalid_count
            }
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Unexpected error during import preview for '{original_filename}': {exc}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Unable to analyze spreadsheet '{original_filename}'. Detail: {str(exc)}"
        )

@router.post("/confirm", response_model=ImportResultResponse, summary="Confirm & Execute Batch Import")
def confirm_import(
    req: ImportConfirmRequest,
    ctx: TenantContext = Depends(require_any_permission(Permission.CONTACTS_MANAGE, Permission.CANDIDATES_MANAGE)),
    db: Session = Depends(get_db)
):
    saved_path = os.path.join(UPLOAD_TMP_DIR, req.file_token)
    if not os.path.exists(saved_path):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uploaded file session expired or not found. Please re-upload.")

    ext = os.path.splitext(req.file_token)[1].lower()
    headers: List[str] = []
    rows_data: List[Dict[str, Any]] = []

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(saved_path, data_only=True)
        sheet = wb[req.sheet_name] if req.sheet_name in wb.sheetnames else wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if all_rows:
            raw_header_row = all_rows[0]
            headers = [
                str(cell).strip() if (cell is not None and str(cell).strip() != "") else f"Unnamed_Col_{col_idx + 1}"
                for col_idx, cell in enumerate(raw_header_row)
            ]
            for row in all_rows[1:]:
                if row and any(val is not None and str(val).strip() != "" for val in row):
                    row_dict = {}
                    for col_idx, val in enumerate(row):
                        if col_idx < len(headers):
                            row_dict[headers[col_idx]] = str(val).strip() if val is not None else ""
                    rows_data.append(row_dict)
    else:  # .csv
        with open(saved_path, "r", encoding="utf-8-sig", errors="ignore") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
            if all_rows:
                raw_header_row = all_rows[0]
                headers = [
                    str(cell).strip() if str(cell).strip() != "" else f"Unnamed_Col_{col_idx + 1}"
                    for col_idx, cell in enumerate(raw_header_row)
                ]
                for row in all_rows[1:]:
                    if row and any(str(val).strip() != "" for val in row):
                        row_dict = {}
                        for col_idx, val in enumerate(row):
                            if col_idx < len(headers):
                                row_dict[headers[col_idx]] = str(val).strip()
                        rows_data.append(row_dict)

    if not rows_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No rows found in uploaded file.")

    imported_contacts = 0
    imported_companies = 0
    reused_companies = 0
    imported_candidates = 0
    updated_records = 0
    skipped_duplicates = 0
    invalid_rows = 0
    errors_count = 0

    try:
        if req.import_type == "CANDIDATES":
            for row in rows_data:
                mapped = {}
                for orig_header, std_field in req.mapping.items():
                    if std_field != "ignore" and orig_header in row:
                        mapped[std_field] = row[orig_header]

                cand_name = format_clean_name(mapped.get("name"))
                email = normalize_email(mapped.get("email"))
                phone = mapped.get("phone")

                if not cand_name and not email and not phone:
                    invalid_rows += 1
                    continue

                clean_email = email.strip().lower() if email else None
                norm_phone = normalize_phone(phone)

                existing = None
                if clean_email:
                    existing = db.query(Candidate).filter(
                        Candidate.organization_id == ctx.organization.id,
                        Candidate.email == clean_email
                    ).first()

                if not existing and norm_phone:
                    existing = db.query(Candidate).filter(
                        Candidate.organization_id == ctx.organization.id,
                        Candidate.phone == phone
                    ).first()

                if existing:
                    if req.duplicate_handling == "SKIP":
                        skipped_duplicates += 1
                        continue
                    elif req.duplicate_handling == "UPDATE":
                        if cand_name: existing.full_name = cand_name
                        if phone: existing.phone = phone
                        if mapped.get("location"): existing.location = mapped.get("location")
                        if mapped.get("primary_skills"): existing.primary_skills = mapped.get("primary_skills")
                        updated_records += 1
                        continue

                exp_val = None
                if mapped.get("experience_years"):
                    try:
                        exp_val = float(mapped["experience_years"])
                    except ValueError:
                        exp_val = 0.0

                candidate = Candidate(
                    organization_id=ctx.organization.id,
                    full_name=cand_name or "Imported Candidate",
                    email=email or f"cand_{uuid.uuid4().hex[:8]}@imported.com",
                    phone=phone,
                    location=mapped.get("location"),
                    primary_skills=mapped.get("primary_skills"),
                    experience_years=exp_val,
                    current_company=mapped.get("current_company"),
                    status=CandidateStatus.NEW,
                    notes="Imported via spreadsheet batch import."
                )
                db.add(candidate)
                imported_candidates += 1
        else:  # HR_CONTACTS
            company_cache: Dict[str, Company] = {}
            existing_companies = db.query(Company).filter(Company.organization_id == ctx.organization.id).all()
            for comp in existing_companies:
                norm_key = normalize_company_name(comp.name)
                if norm_key:
                    company_cache[norm_key] = comp

            for row in rows_data:
                mapped = {}
                for orig_header, std_field in req.mapping.items():
                    if std_field != "ignore" and orig_header in row:
                        mapped[std_field] = row[orig_header]

                person_name = format_clean_name(mapped.get("name"))
                comp_name = mapped.get("company_name")
                email = normalize_email(mapped.get("email"))
                phone = mapped.get("phone")

                if not person_name and not email and not phone:
                    invalid_rows += 1
                    continue

                clean_email = email.strip().lower() if email else None
                norm_phone = normalize_phone(phone)
                norm_person = normalize_person_name(person_name)
                norm_comp = normalize_company_name(comp_name)

                existing = None
                if clean_email:
                    existing = db.query(Contact).filter(
                        Contact.organization_id == ctx.organization.id,
                        Contact.email == clean_email
                    ).first()

                if not existing and norm_phone:
                    existing = db.query(Contact).filter(
                        Contact.organization_id == ctx.organization.id,
                        Contact.phone == phone
                    ).first()

                if not existing and norm_person and norm_comp:
                    existing = (
                        db.query(Contact)
                        .join(Company, Contact.company_id == Company.id)
                        .filter(
                            Contact.organization_id == ctx.organization.id,
                            Contact.name.ilike(person_name)
                        )
                        .first()
                    )

                if existing:
                    if req.duplicate_handling == "SKIP":
                        skipped_duplicates += 1
                        continue
                    elif req.duplicate_handling == "UPDATE":
                        if person_name: existing.name = person_name
                        if mapped.get("designation"): existing.designation = mapped.get("designation")
                        if phone: existing.phone = phone
                        if email: existing.email = email
                        if mapped.get("linkedin_url"): existing.linkedin_url = mapped.get("linkedin_url")
                        if mapped.get("location"): existing.location = mapped.get("location")
                        if mapped.get("notes"): existing.notes = mapped.get("notes")
                        updated_records += 1
                        continue

                target_company_id = None
                if comp_name:
                    norm_comp_key = normalize_company_name(comp_name)
                    if norm_comp_key and norm_comp_key in company_cache:
                        target_company_id = company_cache[norm_comp_key].id
                        reused_companies += 1
                    else:
                        new_comp = Company(
                            organization_id=ctx.organization.id,
                            name=comp_name.strip(),
                            location=mapped.get("location")
                        )
                        db.add(new_comp)
                        db.flush()
                        if norm_comp_key:
                            company_cache[norm_comp_key] = new_comp
                        target_company_id = new_comp.id
                        imported_companies += 1

                timestamp_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                contact = Contact(
                    organization_id=ctx.organization.id,
                    company_id=target_company_id,
                    name=person_name or "HR Contact",
                    designation=mapped.get("designation"),
                    phone=phone,
                    email=email,
                    linkedin_url=mapped.get("linkedin_url"),
                    location=mapped.get("location"),
                    source=f"Spreadsheet Import ({timestamp_str})",
                    status=ContactStatus.NOT_CONTACTED,
                    notes=mapped.get("notes")
                )
                db.add(contact)
                imported_contacts += 1

        db.commit()

        if os.path.exists(saved_path):
            os.remove(saved_path)

        record_label = "candidates" if req.import_type == "CANDIDATES" else "HR contacts"
        imported_count = imported_candidates if req.import_type == "CANDIDATES" else imported_contacts

        return ImportResultResponse(
            success=True,
            imported_contacts_count=imported_contacts,
            imported_companies_count=imported_companies,
            reused_companies_count=reused_companies,
            imported_candidates_count=imported_candidates,
            updated_records_count=updated_records,
            skipped_duplicates_count=skipped_duplicates,
            invalid_rows_count=invalid_rows,
            errors_count=errors_count,
            message=f"Import successful! Created {imported_count} {record_label}, updated {updated_records}, reused {reused_companies} companies."
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        logger.error(f"Import confirm execution failed for token '{req.file_token}': {e}", exc_info=True)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Import execution failed and transaction was safely rolled back. Detail: {str(e)}"
        )

